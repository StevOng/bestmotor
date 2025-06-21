import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.db.models import Sum, Q
from ...decorators import *
from ...models.barang import *

@both_required
def barang(request):
    filter = request.GET.get('filter')
    dari = request.GET.get('dari_tgl')
    sampe = request.GET.get('smpe_tgl')

    barang = Barang.objects.none()

    if filter == "laku" and dari and sampe:
        barang = get_barang_laku(dari, sampe)
    elif filter == "rendah":
        barang = Barang.objects.filter(stok__lt=models.F('stok_minimum'))
    else:
        barang = Barang.objects.prefetch_related('detailpesanan_set').all()

    for brg in barang:
        brg.total_pesanan = sum(d.qty_pesan for d in brg.detailpesanan_set.all())
        if brg.stok <= brg.stok_minimum:
            brg.selisih = brg.stok_minimum - brg.stok
        else:
            brg.selisih = 0

    return render(request, 'barang/barang.html', {'barang': barang,'filter': filter})

@admin_required
def tambah_barang(request, id=None):
    barang = None
    tier_harga = []

    if id:
        barang = Barang.objects.get(id=id)
        tier_harga = TierHarga.objects.filter(barang_id=barang.id)
    return render(request, 'barang/tambahbrg.html', {'detail_barang': barang, 'tier_harga': tier_harga})

def get_barang_laku(dari, sampe):
    dari = parse_date(dari)
    sampe = parse_date(sampe)

    return Barang.objects.annotate(total_terjual=Sum(
        'detailpesanan_qty_pesan',
        filter=Q(detailpesanan__pesanan_id__tanggal_pesanan__range=[dari,sampe])
    )).filter(total_terjual__gt=0)

def export_excel(request):
    work_book = openpyxl.Workbook()

    # ===== SHEET 1: List Barang =====
    sheet_barang = work_book.active
    sheet_barang.title = "List Barang"

    col_heads_barang = [
        'No',
        'Kode Barang',
        'Nama Barang',
        'Tipe Motor',
        'Merk',
        'Harga Jual',
        'Stok Minimum',
        'Harga Modal',
        'Stok',
        'Qty Terjual',
        'Keterangan'
    ]
    sheet_barang.append(col_heads_barang)

    header_font = Font(bold=True)
    for col_num, column_title in enumerate(col_heads_barang, start=1):
        sheet_barang.cell(row=1, column=col_num).font = header_font

    for idx, barang in enumerate(Barang.objects.all(), start=1):
        sheet_barang.append([
            idx,
            barang.kode_barang,
            barang.nama_barang,
            barang.tipe,
            barang.merk,
            barang.harga_jual,
            barang.stok_minimum,
            barang.harga_modal,
            barang.stok,
            barang.qty_terjual,
            barang.keterangan
        ])

    for col in sheet_barang.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet_barang.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # ===== SHEET 2: Tier Harga =====
    sheet_tier = work_book.create_sheet(title="Tier Harga")

    col_heads_tier = [
        'No',
        'Kode Barang',
        'Nama Barang',
        'Min Qty Grosir',
        'Harga Satuan'
    ]
    sheet_tier.append(col_heads_tier)
    for col_num, column_title in enumerate(col_heads_tier, start=1):
        sheet_tier.cell(row=1, column=col_num).font = header_font

    for idx, tier in enumerate(TierHarga.objects.select_related('barang_id'), start=1):
        sheet_tier.append([
            idx,
            tier.barang.kode_barang,
            tier.barang.nama_barang,
            tier.min_qty_grosir,
            tier.harga_satuan
        ])

    for col in sheet_tier.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet_tier.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # ===== Simpan & Response =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=barang_dan_tier.xlsx'
    work_book.save(response)

    return response